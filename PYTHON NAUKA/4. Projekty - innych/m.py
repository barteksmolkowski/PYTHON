import os

import openpyxl
from openpyxl import Workbook

DATA_FILE = "xd_data.xlsx"

class zlaWartoscDlaTypException(Exception):
    def __init__(self):
        tekst = "Podałeś złą wartość"
        super().__init__(tekst)

def autoZapis(func):
    def wrapper(wb, *args, **kwargs):
        wynik = func(wb, *args, **kwargs)
        wb.save(DATA_FILE)
        print("Zmiany zapisane do pliku.")
        return wynik
    return wrapper

def zArkuszem(func):
    def wrapper(wb, *args, **kwargs):
        ws = wb["Sprzedaz"]
        return func(ws, *args, **kwargs)
    return wrapper

def is_valid_record(ws, row_number):
    max_row = ws.max_row - 1
    return 1 <= row_number <= max_row

def load_workbook():
    if not os.path.exists(DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sprzedaz"
        ws.append(["Produkt", "Cena zakupu", "Cena sprzedazy", "Ilosc", "Prowizja", "Zysk jednostkowy", "Zysk calkowity"])
        wb.save(DATA_FILE)
    else:
        wb = openpyxl.load_workbook(DATA_FILE)
    return wb

@zArkuszem
def save_record(ws, produkt, zakup, sprzedaz, ilosc, prowizja):
    ws.append([produkt, zakup, sprzedaz, ilosc, prowizja, zakup - sprzedaz - prowizja, (zakup - sprzedaz - prowizja) * ilosc])
    print(" Rekord zapisany!")

def total_profit(wb):
    print(f"\n Laczny zysk: {sum(row[6] for row in wb["Sprzedaz"].iter_rows(min_row=2, values_only=True))} zl")

@autoZapis
@zArkuszem
def update_record(ws, row_number, zakup=None, sprzedaz=None, ilosc=None, prowizja=None):
    if not is_valid_record(ws, row_number):
        print(f"Niepoprawny numer rekordu! Powinien być od 1 do {ws.max_row - 1}.")
        return

    row = list(ws.iter_rows(min_row=row_number + 1, max_row=row_number + 1, values_only=True))[0]
    produkt, zakup0, sprzedaz0, ilosc0, prowizja0, *_ = row

    zakup = zakup if zakup is not None else zakup0
    sprzedaz = sprzedaz if sprzedaz is not None else sprzedaz0
    ilosc = ilosc if ilosc is not None else ilosc0
    prowizja = prowizja if prowizja is not None else prowizja0

    for col, value in enumerate([produkt, zakup, sprzedaz, ilosc, prowizja, sprzedaz - zakup - prowizja, (sprzedaz - zakup - prowizja) * ilosc], start=1):
        ws.cell(row=row_number + 1, column=col).value = value

    print("Rekord zaktualizowany!")

@autoZapis
@zArkuszem
def delete_record(ws, row_number):    
    if isinstance(row_number, str) and row_number.lower() == "all":
        max_row = ws.max_row
        if max_row > 1:
            ws.delete_rows(2, max_row - 1)
            print("Wszystkie rekordy zostały usunięte!")
        else:
            print("Brak rekordów do usunięcia.")
        return

    if not isinstance(row_number, int):
        print("Podaj poprawny numer rekordu lub wpisz 'all'.")
        return

    max_row = ws.max_row - 1
    if not (1 <= row_number <= max_row):
        print(f"Niepoprawny numer rekordu! Powinien być od 1 do {max_row}.")
        return

    ws.delete_rows(row_number + 1)
    print(f"Rekord nr {row_number} usunięty!")

def pobierz_pola(pola: list | str, wb):
    pokaz_wszystkie_rekordy(wb, False)
    definicje = {
        "produkt": ("Nazwa produktu", str, False, None),
        "zakup": ("Cena zakupu", float, False, None),
        "sprzedaz": ("Cena sprzedaży", float, False, None),
        "ilosc": ("Ilość sztuk", int, False, None),
        "prowizja_input": ("Prowizja (ENTER = 0)", float, True, 0.0),
        "n1": ("Który rekord zaktualizować (numer) [ENTER = bez zmian]\nTwoja zmiana: ", int, True, None),
        "zakupN": ("Nowa cena zakupu", float, True, None),
        "sprzedazN": ("Nowa cena sprzedaży", float, True, None),
        "iloscN": ("Nowa ilość", int, True, None),
        "prowizjaN": ("Nowa prowizja", float, True, None),
        "n2": ("Który rekord usunąć (numer)", int, False, None),
    }

    wyniki = {}

    for pole in [pola] if isinstance(pola, str) else pola:
        if pole not in definicje:
            raise ValueError(f"Pole '{pole}' nie istnieje w definicjach!")

        opis, typ, pozwol_none, default = definicje[pole]

        valid = False
        while not valid:
            raw = input(f"{opis}: ")

            if raw == "":
                if default is not None:
                    wyniki[pole], valid = default, True
                elif pozwol_none:
                    wyniki[pole], valid = None, True
                else:
                    print("To pole jest wymagane.")
            else:
                try:
                    wyniki[pole], valid = typ(raw), True
                except ValueError:
                    print(f"Podaj wartość typu {typ.__name__}.")

    return tuple(wyniki[p] for p in pola)

@zArkuszem
def pokaz_wszystkie_rekordy(ws, pustePokazac = True):
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if pustePokazac:
        print("\nLista rekordów:")
        if not rows:
            print("Brak rekordów do wyświetlenia.")
        [print("-"*20+f"\nProdukt nr.{idx}:\n# 1 # Nazwa: {row[0]}\n# 2 # Sztuk: {row[3]}\n# 3 # Prowizja: {row[4]}\n# 4 # Koszt: {row[1]}zł\n# 5 # Sprzedaż: {row[2]}zł\n# 6 # Zysk jednostkowy: {row[5]}zł\n# 7 # Końcowy zarobek: {row[6]}zł") for idx, row in enumerate(rows, start=1)]; print("-"*20)

def main():
    wb = load_workbook()

    slownik = {
        1: lambda: save_record(wb, *pobierz_pola(["produkt", "zakup", "sprzedaz", "ilosc", "prowizja_input"], wb)),
        2: lambda: pokaz_wszystkie_rekordy(wb),
        3: lambda: total_profit(wb),
        4: lambda: update_record(wb, *pobierz_pola(["n1", "zakupN", "sprzedazN", "iloscN", "prowizjaN"], wb)),
        5: lambda: delete_record(wb, *pobierz_pola(["n2"], wb)),
        6: lambda: print("Do zobaczenia!")
    }

    while True:
        print("\nMenu:")
        print("1. Dodaj rekord")
        print("2. Pokaż wszystkie rekordy")
        print("3. Pokaż łączny zysk")
        print("4. Aktualizuj rekord")
        print("5. Usuń rekord")
        print("6. Wyjście")

        try:
            wybor = int(input("Wybierz opcję: "))
            if wybor == 6:
                slownik[6]()
                break
            elif wybor in slownik:
                slownik[wybor]()
            else:
                print("Zły wybór!")
        except ValueError:
            print("Podaj poprawny numer opcji!")

main()